#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import sys
import argparse
import subprocess
import os
import platform
import tempfile
from pathlib import Path
from openpyxl import load_workbook


def _default_libreoffice_user_profile_dir():
    if platform.system() == "Darwin":
        # Default LibreOffice user profile location on macOS.
        return os.path.expanduser("~/Library/Application Support/LibreOffice/4")
    # Linux default profile.
    return os.path.expanduser("~/.config/libreoffice/4")


def _macro_dir_for_profile(profile_dir: str) -> str:
    # LibreOffice profile contains a `user/` subtree.
    return os.path.join(profile_dir, "user", "basic", "Standard")


def setup_libreoffice_macro(profile_dir: str) -> bool:
    """Setup LibreOffice macro for recalculation inside the given profile dir."""
    macro_dir = _macro_dir_for_profile(profile_dir)
    macro_file = os.path.join(macro_dir, "Module1.xba")

    if os.path.exists(macro_file):
        with open(macro_file, "r") as f:
            if "RecalculateAndSave" in f.read():
                return True

    if not os.path.exists(macro_dir):
        os.makedirs(macro_dir, exist_ok=True)

    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def _file_url(path: str) -> str:
    return Path(path).absolute().as_uri()


def recalc(filename: str, timeout: int = 30, *, isolated: bool = True, keep_profile: bool = False):
    """
    Recalculate formulas in Excel file and report any errors
    
    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)
    
    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}
    
    abs_path = str(Path(filename).absolute())

    # If isolated is enabled, create a temporary LibreOffice profile directory so we
    # don't permanently write macros into the user's real LibreOffice profile.
    temp_profile_ctx = None
    profile_dir = _default_libreoffice_user_profile_dir()
    if isolated:
        if keep_profile:
            profile_dir = tempfile.mkdtemp(prefix="xlsx-recalc-lo-profile-")
        else:
            temp_profile_ctx = tempfile.TemporaryDirectory(prefix="xlsx-recalc-lo-profile-")
            profile_dir = temp_profile_ctx.name

    try:
        if not setup_libreoffice_macro(profile_dir):
            return {'error': 'Failed to setup LibreOffice macro'}

        cmd = [
            'soffice',
            '--headless',
            '--norestore',
        ]
        if isolated:
            cmd.append(f'--env:UserInstallation={_file_url(profile_dir)}')
        cmd += [
            'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
            abs_path,
        ]
    
        # Handle timeout command differences between Linux and macOS
        if platform.system() != 'Windows':
            timeout_cmd = 'timeout' if platform.system() == 'Linux' else None
            if platform.system() == 'Darwin':
                # Check if gtimeout is available on macOS
                try:
                    subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                    timeout_cmd = 'gtimeout'
                except (FileNotFoundError, subprocess.TimeoutExpired):
                    pass

            if timeout_cmd:
                cmd = [timeout_cmd, str(timeout)] + cmd

        result = subprocess.run(cmd, capture_output=True, text=True)
    
        if result.returncode != 0 and result.returncode != 124:  # 124 is timeout exit code
            error_msg = result.stderr or 'Unknown error during recalculation'
            if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
                return {'error': 'LibreOffice macro not configured properly'}
            else:
                return {'error': error_msg}
    
        # Check for Excel errors in the recalculated file - scan ALL cells
        try:
            wb = load_workbook(filename, data_only=True)
        
            excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
            error_details = {err: [] for err in excel_errors}
            total_errors = 0
        
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Check ALL rows and columns - no limits
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            for err in excel_errors:
                                if err in cell.value:
                                    location = f"{sheet_name}!{cell.coordinate}"
                                    error_details[err].append(location)
                                    total_errors += 1
                                    break
        
            wb.close()
        
            # Build result summary
            result = {
                'status': 'success' if total_errors == 0 else 'errors_found',
                'total_errors': total_errors,
                'error_summary': {},
                'isolated_profile': bool(isolated),
            }
        
            # Add non-empty error categories
            for err_type, locations in error_details.items():
                if locations:
                    result['error_summary'][err_type] = {
                        'count': len(locations),
                        'locations': locations[:20]  # Show up to 20 locations
                    }
        
            # Add formula count for context - also check ALL cells
            wb_formulas = load_workbook(filename, data_only=False)
            formula_count = 0
            for sheet_name in wb_formulas.sheetnames:
                ws = wb_formulas[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_count += 1
            wb_formulas.close()
        
            result['total_formulas'] = formula_count
        
            # If we used an isolated profile and the user wants to keep it, surface the path.
            if isolated and keep_profile:
                result['profile_dir'] = profile_dir

            return result
        
        except Exception as e:
            return {'error': str(e)}
    finally:
        if isolated and temp_profile_ctx is not None:
            temp_profile_ctx.cleanup()


def main():
    parser = argparse.ArgumentParser(
        description="Recalculate all formulas in an Excel file using LibreOffice (with optional isolated profile).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("excel_file", help="Path to Excel file (.xlsx)")
    parser.add_argument("timeout_seconds", nargs="?", type=int, default=30, help="Timeout in seconds")
    parser.add_argument(
        "--no-isolated",
        action="store_true",
        help="Use the default LibreOffice user profile (may write a macro into your real profile)",
    )
    parser.add_argument(
        "--keep-profile",
        action="store_true",
        help="Keep the temporary LibreOffice profile directory (useful for debugging).",
    )

    args = parser.parse_args()

    result = recalc(
        args.excel_file,
        args.timeout_seconds,
        isolated=not args.no_isolated,
        keep_profile=args.keep_profile,
    )
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
